from rest_framework import viewsets, permissions, status
from rest_framework.response import Response
from rest_framework.decorators import action
from django.utils import timezone
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import PayrollPeriod, PayrollAdjustment, Holiday, Payroll
from .serializers import (
    PayrollPeriodSerializer, PayrollAdjustmentSerializer,
    HolidaySerializer, PayrollSerializer
)
from .services import process_payroll_for_employee
from employees.models import Employee
from employees.permissions import IsPayrollOrAdmin, IsHROrAdmin

class PayrollPeriodViewSet(viewsets.ModelViewSet):
    queryset = PayrollPeriod.objects.all().order_by('-start_date')
    serializer_class = PayrollPeriodSerializer
    permission_classes = [IsPayrollOrAdmin]

    @action(detail=True, methods=['post'], url_path='process')
    def process_payroll(self, request, pk=None):
        period = self.get_object()
        if period.is_locked:
            return Response(
                {"detail": "This payroll period is locked and cannot be processed/modified."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get all active employees
        employees = Employee.objects.filter(user__is_active=True)
        payrolls = []
        for emp in employees:
            p = process_payroll_for_employee(emp, period, request.user)
            payrolls.append(p)
            
        return Response({
            "detail": f"Successfully processed payroll for {len(payrolls)} employees.",
            "processed_count": len(payrolls)
        })

    @action(detail=True, methods=['post'], url_path='lock')
    def lock_period(self, request, pk=None):
        period = self.get_object()
        if period.is_locked:
            return Response({"detail": "Period is already locked."}, status=status.HTTP_400_BAD_REQUEST)
        
        # Lock
        period.is_locked = True
        period.locked_at = timezone.now()
        period.locked_by = request.user
        period.save()
        
        # Generate final payslips for this period (we can trigger this here or separately)
        # We will check if payslips are created inside payslips app
        from payslips.models import Payslip
        from payslips.services import generate_pdf_payslip
        
        payrolls = Payroll.objects.filter(period=period)
        payslip_count = 0
        for pr in payrolls:
            # Generate code: PS-YYYY-MM-EMPID
            ps_code = f"PS-{period.start_date.year}-{period.start_date.month:02d}-{pr.employee.employee_code}"
            payslip, created = Payslip.objects.update_or_create(
                payroll=pr,
                defaults={'payslip_code': ps_code}
            )
            # Try to build PDF
            try:
                generate_pdf_payslip(payslip)
                payslip_count += 1
            except Exception as e:
                print(f"Failed to generate PDF for {ps_code}:", e)

        return Response({
            "detail": f"Payroll period locked. Generated {payslip_count} final payslip documents.",
            "is_locked": True
        })

    @action(detail=True, methods=['get'], url_path='salary-register')
    def export_salary_register(self, request, pk=None):
        period = self.get_object()
        payrolls = Payroll.objects.filter(period=period).select_related('employee', 'employee__user', 'employee__department', 'employee__designation')
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Salary Register"
        
        # Enable grid lines
        ws.views.sheetView[0].showGridLines = True
        
        # Styling definitions
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid") # Slate 900
        title_font = Font(name="Segoe UI", size=16, bold=True, color="0F172A")
        sub_font = Font(name="Segoe UI", size=11, italic=True)
        data_font = Font(name="Segoe UI", size=10)
        total_font = Font(name="Segoe UI", size=11, bold=True)
        total_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid") # Slate 100
        
        border_thin = Side(style='thin', color='CBD5E1')
        data_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
        border_double = Side(style='double', color='0F172A')
        total_border = Border(top=border_thin, bottom=border_double)
        
        # Add Title Info
        ws.append([])
        ws.cell(row=2, column=2, value="Staff Payroll Management System").font = title_font
        ws.cell(row=3, column=2, value=f"Salary Register - {period.name} ({period.start_date} to {period.end_date})").font = sub_font
        ws.append([])
        ws.append([])
        
        # Headers
        headers = [
            "Emp Code", "Employee Name", "Department", "Designation",
            "Basic Salary", "HRA", "DA", "Travel Allow.", "Medical Allow.", 
            "Bonus/Allow.", "Overtime Pay", "Gross Earnings",
            "PF (12%)", "ESI (0.75%)", "Prof. Tax", "Income Tax (TDS)",
            "Leave Deduct", "Late Penalty", "Gross Deduct", "Net Salary"
        ]
        
        ws.append(headers)
        header_row = 6
        
        # Style headers
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = data_border
        
        ws.row_dimensions[header_row].height = 28
        
        start_row = 7
        row_idx = start_row
        
        # Add Data
        for pr in payrolls:
            basic = float(pr.basic_salary)
            hra = float(pr.hra)
            da = float(pr.da)
            travel = float(pr.travel_allowance)
            medical = float(pr.medical_allowance)
            bonus = float(pr.bonus)
            ot = float(pr.overtime_pay)
            
            pf = float(pr.pf_deduction)
            esi = float(pr.esi_deduction)
            pt = float(pr.professional_tax)
            tds = float(pr.income_tax)
            leave_ded = float(pr.leave_deductions)
            late = float(pr.late_penalties)
            other_adj = float(pr.other_adjustments)
            
            # Gross values using excel formulas for columns
            # Column L (Gross Earnings) = E + F + G + H + I + J + K
            # Column S (Gross Deductions) = M + N + O + P + Q + R
            # Column T (Net Salary) = L - S
            
            ws.append([
                pr.employee.employee_code,
                pr.employee.user.get_full_name(),
                pr.employee.department.name if pr.employee.department else "-",
                pr.employee.designation.title if pr.employee.designation else "-",
                basic, hra, da, travel, medical, bonus, ot,
                f"=SUM(E{row_idx}:K{row_idx})", # L: Gross Earnings
                pf, esi, pt, tds, leave_ded, late,
                f"=SUM(M{row_idx}:R{row_idx})", # S: Gross Deductions
                f"=L{row_idx}-S{row_idx}"      # T: Net Salary
            ])
            
            # Format row
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.border = data_border
                
                # Alignments and number formats
                if col_idx in [1, 3, 4]:
                    cell.alignment = Alignment(horizontal="center")
                elif col_idx > 4:
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = "$#,##0.00"
            
            row_idx += 1
            
        end_row = row_idx - 1
        
        # Add Totals Row
        ws.append([
            "TOTAL", "", "", "",
            f"=SUM(E{start_row}:E{end_row})",
            f"=SUM(F{start_row}:F{end_row})",
            f"=SUM(G{start_row}:G{end_row})",
            f"=SUM(H{start_row}:H{end_row})",
            f"=SUM(I{start_row}:I{end_row})",
            f"=SUM(J{start_row}:J{end_row})",
            f"=SUM(K{start_row}:K{end_row})",
            f"=SUM(L{start_row}:L{end_row})",
            f"=SUM(M{start_row}:M{end_row})",
            f"=SUM(N{start_row}:N{end_row})",
            f"=SUM(O{start_row}:O{end_row})",
            f"=SUM(P{start_row}:P{end_row})",
            f"=SUM(Q{start_row}:Q{end_row})",
            f"=SUM(R{start_row}:R{end_row})",
            f"=SUM(S{start_row}:S{end_row})",
            f"=SUM(T{start_row}:T{end_row})"
        ])
        
        total_row = row_idx
        ws.row_dimensions[total_row].height = 22
        
        # Style Totals Row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=total_row, column=col_idx)
            cell.font = total_font
            cell.fill = total_fill
            cell.border = total_border
            if col_idx > 4:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "$#,##0.00"
            else:
                cell.alignment = Alignment(horizontal="left")
                
        # Auto-fit columns
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if cell.row < 5:  # skip title cells for size calc
                    continue
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 11)
            
        # Return response
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = f'attachment; filename="salary_register_{period.name.replace(" ", "_")}.xlsx"'
        wb.save(response)
        return response

class PayrollAdjustmentViewSet(viewsets.ModelViewSet):
    queryset = PayrollAdjustment.objects.all()
    serializer_class = PayrollAdjustmentSerializer
    permission_classes = [IsPayrollOrAdmin]

    def get_queryset(self):
        user = self.request.user
        if user.role in ['HRAdmin', 'SuperAdmin', 'PayrollOfficer']:
            return self.queryset
        # Employees cannot view adjustments directly
        return self.queryset.none()

class HolidayViewSet(viewsets.ModelViewSet):
    queryset = Holiday.objects.all().order_by('date')
    serializer_class = HolidaySerializer
    permission_classes = [IsHROrAdmin]

class PayrollViewSet(viewsets.ModelViewSet):
    queryset = Payroll.objects.select_related('employee', 'period').all()
    serializer_class = PayrollSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.role in ['HRAdmin', 'SuperAdmin', 'PayrollOfficer']:
            return self.queryset
        # Regular employee
        return self.queryset.filter(employee__user=user)
