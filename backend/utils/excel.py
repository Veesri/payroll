import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

def generate_excel_report(title, headers, data_rows):
    """
    Generates an Excel workbook buffer with a formatted title, headers, and auto-sized columns.
    data_rows should be a list of lists/tuples corresponding to the headers.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    
    # Fonts & Styles
    title_font = Font(size=16, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid") # Indigo theme
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid") # Light gray
    center_align = Alignment(horizontal="center", vertical="center")
    
    # Add Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = center_align
    
    # Add Generated Date
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    date_cell = ws.cell(row=2, column=1, value=f"Generated On: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    date_cell.alignment = Alignment(horizontal="right")
    
    # Empty Row
    ws.append([])
    
    # Add Headers
    ws.append(headers)
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=4, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        
    # Add Data
    for row in data_rows:
        # Convert any None values to empty strings to avoid openpyxl issues
        cleaned_row = [str(item) if item is not None else "" for item in row]
        ws.append(cleaned_row)
        
    # Auto-size columns
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column) # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        # Limit max width to avoid crazy wide columns
        if adjusted_width > 50:
            adjusted_width = 50
        ws.column_dimensions[column].width = adjusted_width
        
    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output
